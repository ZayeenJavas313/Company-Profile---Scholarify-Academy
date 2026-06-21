from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAdminUser
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.utils import timezone
from django.db.models import Count, Avg, Q
from datetime import datetime
from .models import Subtest, Soal, HasilTryout, Batch
from django.http import JsonResponse

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@api_view(['GET'])
def subtests_list(request):
    """
    GET /api/subtests/
    Returns list of all subtests with their details.
    """
    subtests = Subtest.objects.all().order_by('code')
    data = []
    for subtest in subtests:
        data.append({
            'id': subtest.code.lower(),
            'code': subtest.code,
            'title': subtest.nama_subtest,
            'description': f"Subtest {subtest.nama_subtest} untuk UTBK 2025",
            'duration': float(subtest.durasi_menit),
            'questionCount': subtest.jumlah_soal,
        })
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
def subtest_questions(request, code):
    """
    GET /api/subtests/{code}/questions/?batch_id=batch-1
    Returns all questions for a specific subtest, optionally filtered by batch_id.
    """
    try:
        subtest = Subtest.objects.get(code=code.upper())
    except Subtest.DoesNotExist:
        return Response(
            {'error': f'Subtest with code {code} not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Get batch_id from query parameters
    batch_id = request.GET.get('batch_id', '').strip()
    
    # Filter questions by subtest and optionally by batch
    if batch_id:
        try:
            # Filter by batch_id
            questions = Soal.objects.filter(
                subtest=subtest,
                batch__batch_id=batch_id
            ).select_related('batch').order_by('id')
            
            # Log for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Filtering questions for subtest {code} by batch_id='{batch_id}', found {questions.count()} question(s)")
            
            # If no questions found for this batch, log warning but still return empty list
            if questions.count() == 0:
                logger.warning(f"No questions found for subtest {code} and batch_id '{batch_id}'")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error filtering questions by batch_id: {e}", exc_info=True)
            # Fallback: filter by subtest only
            questions = Soal.objects.filter(subtest=subtest).order_by('id')
    else:
        # No batch_id provided, return all questions for this subtest
        questions = Soal.objects.filter(subtest=subtest).order_by('id')
    data = []
    for idx, soal in enumerate(questions, start=1):
        # URL gambar jika ada - gunakan path relatif /media/ agar frontend bisa resolve dengan BACKEND_ORIGIN
        image_url = None
        if soal.soal_image:
            # Gunakan path relatif /media/ yang akan di-resolve oleh frontend
            media_path = str(soal.soal_image.url) if hasattr(soal.soal_image, 'url') else ''
            # Pastikan path dimulai dengan /media/
            if media_path and not media_path.startswith('/'):
                media_path = '/' + media_path
            image_url = media_path
        
        data.append({
            'id': f"{subtest.code.lower()}-{idx}",
            'soal_id': soal.id,  # ID sebenarnya dari database
            'subtestId': subtest.code.lower(),
            'question': soal.soal_text,
            'question_image': image_url,  # Path relatif gambar soal (jika ada)
            'options': [
                {'key': 'A', 'text': soal.option_a},
                {'key': 'B', 'text': soal.option_b},
                {'key': 'C', 'text': soal.option_c},
                {'key': 'D', 'text': soal.option_d},
                {'key': 'E', 'text': soal.option_e},
            ],
            'answer': soal.correct_answer,
        })
    return Response(data, status=status.HTTP_200_OK)


@api_view(['POST'])
@csrf_exempt
@ensure_csrf_cookie
def login_view(request):
    """
    POST /api/auth/login/

    Body: { "username": "...", "password": "..." }
    Validasi user menggunakan database Django default.
    Membuat Django session pada user login sehingga POST requests bisa autentikasi via sessionid cookie.
    """
    import json
    import logging
    # from django.http import JsonResponse # Tidak perlu lagi, pakai Response dari DRF
    
    logger = logging.getLogger(__name__)
    logger.info("=== login_view called ===")
    logger.info(f"Content-Type: {request.content_type}")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Request body (first 200 chars): {str(request.body)[:200]}")
    
    # Parse JSON body - DRF request.data sudah handle parsing untuk Content-Type: application/json
    try:
        # DRF's request.data handles JSON parsing automatically
        if hasattr(request, 'data') and request.data:
            data = request.data
            logger.info(f"Using request.data: {data}")
        elif request.body:
            # Fallback: manual parse jika request.data tidak ada
            data = json.loads(request.body.decode('utf-8'))
            logger.info(f"Manually parsed body: {data}")
        else:
            data = {}
            logger.warning("No request body found")
    except (json.JSONDecodeError, ValueError, AttributeError, UnicodeDecodeError) as e:
        logger.error(f"Error parsing request body: {e}", exc_info=True)
        return Response({'error': 'Invalid JSON'}, status=status.HTTP_400_BAD_REQUEST)
    
    username = data.get('username', '').strip() if isinstance(data, dict) else ''
    password = data.get('password', '') if isinstance(data, dict) else ''
    
    logger.info(f"Extracted username: '{username}' (length: {len(username)})")
    logger.info(f"Extracted password: {'*' * len(password) if password else '(empty)'} (length: {len(password)})")

    if not username or not password:
        return Response(
            {'error': 'Username & password wajib'},
            status=status.HTTP_400_BAD_REQUEST
        )

    user = authenticate(request, username=username, password=password)
    if user is None:
        logger.warning(f"Authentication failed for username: {username}")
        return Response(
            {'error': 'Username atau password salah'},
            status=status.HTTP_401_UNAUTHORIZED
        )
    
    # Cek jika user is_active
    if not user.is_active:
        logger.warning(f"User {username} is not active")
        return Response(
            {'error': 'Akun Anda tidak aktif. Hubungi admin untuk mengaktifkan akun.'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    logger.info(f"User {username} authenticated successfully, is_staff: {user.is_staff}")

    # PENTING: Panggil login() untuk membuat Django session (sessionid cookie)
    login(request, user)

    # Bentuk payload minimal yang akan dipakai frontend
    name = user.get_full_name() or user.username
    role = 'admin' if user.is_staff else 'student'

    # Hapus semua manual sessionid handling
    # Biarkan Django middleware yang menyetel cookie sessionid

    return Response(
        {
            'ok': True,
            'user': {
                'username': user.username,
                'name': name,
                'role': role,
            },
        },
        status=status.HTTP_200_OK
    )


@api_view(['POST'])
@csrf_exempt
def import_soal_excel(request):
    """
    POST /api/import-soal-excel/
    
    Import bank soal dari file Excel.
    Format: Kode Subtes | SOAL | A | B | C | D | E | KUNCI
    
    Body: multipart/form-data dengan field 'file' (file Excel .xlsx)
    """
    if load_workbook is None:
        return Response(
            {'error': 'openpyxl belum terinstall. Install dengan: pip install openpyxl'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    if 'file' not in request.FILES:
        return Response(
            {'error': 'File Excel tidak ditemukan. Kirim dengan field "file".'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    excel_file = request.FILES['file']
    
    # validasi ekstensi
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'File harus berformat .xlsx atau .xls'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # baca file Excel dengan openpyxl
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        
        # baca header
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        expected_headers = ["Kode Subtes", "SOAL", "A", "B", "C", "D", "E", "KUNCI"]
        
        # Cek apakah ada kolom gambar (opsional, kolom ke-9)
        has_gambar_column = len(header_row) > 8 and str(header_row[8]).strip().lower() in ['gambar', 'gambar_soal', 'image', 'gambar soal']
        
        # validasi header (case-insensitive)
        header_lower = [str(h).strip().lower() if h else "" for h in header_row[:8]]
        expected_lower = [h.lower() for h in expected_headers]
        
        if header_lower[:8] != expected_lower:
            return Response(
                {
                    'error': f'Header tidak sesuai. Diharapkan: {", ".join(expected_headers)}',
                    'found': header_lower[:8]
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created = 0
        errors = []
        
        # proses setiap baris data
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # skip baris kosong
            if not any(row[:8]):
                continue
            
            try:
                kode_subtes = str(row[0]).strip() if row[0] else ""
                soal_text = str(row[1]).strip() if row[1] else ""
                option_a = str(row[2]).strip() if row[2] else ""
                option_b = str(row[3]).strip() if row[3] else ""
                option_c = str(row[4]).strip() if row[4] else ""
                option_d = str(row[5]).strip() if row[5] else ""
                option_e = str(row[6]).strip() if row[6] else ""
                kunci = str(row[7]).strip().upper() if row[7] else ""
                gambar_path = str(row[8]).strip() if has_gambar_column and len(row) > 8 and row[8] else ""
                
                # validasi
                if not kode_subtes:
                    errors.append(f"Baris {row_num}: Kode Subtes kosong")
                    continue
                
                # SOAL atau GAMBAR harus ada salah satu
                if not soal_text and not gambar_path:
                    errors.append(f"Baris {row_num}: SOAL atau GAMBAR harus diisi (minimal salah satu)")
                    continue
                
                if kunci not in ["A", "B", "C", "D", "E"]:
                    errors.append(f"Baris {row_num}: KUNCI harus A/B/C/D/E, ditemukan: {kunci}")
                    continue
                
                # cari subtest
                try:
                    subtest = Subtest.objects.get(code=kode_subtes)
                except Subtest.DoesNotExist:
                    errors.append(f"Baris {row_num}: Kode Subtes '{kode_subtes}' tidak ditemukan di database")
                    continue
                
                # Handle gambar
                soal_image = None
                if gambar_path:
                    from django.core.files import File
                    from urllib.parse import urlparse
                    try:
                        import requests  # type: ignore
                    except ImportError:
                        requests = None
                    import tempfile
                    import os
                    
                    # Cek apakah path adalah URL
                    parsed = urlparse(gambar_path)
                    if parsed.scheme in ['http', 'https']:
                        # Download dari URL
                        if requests is None:
                            raise ImportError("requests library is required for downloading images from URLs")
                        try:
                            response = requests.get(gambar_path, timeout=10)
                            response.raise_for_status()
                            # Simpan ke temporary file
                            img_ext = os.path.splitext(parsed.path)[1] or '.jpg'
                            with tempfile.NamedTemporaryFile(delete=False, suffix=img_ext) as img_tmp:
                                img_tmp.write(response.content)
                                img_tmp_path = img_tmp.name
                            
                            # Buka file dan simpan ke model
                            with open(img_tmp_path, 'rb') as img_file:
                                soal_image = File(img_file, name=os.path.basename(parsed.path) or f'image_{row_num}{img_ext}')
                            os.unlink(img_tmp_path)
                        except Exception as e:
                            errors.append(f"Baris {row_num}: Gagal download gambar dari URL: {str(e)}")
                    else:
                        # Path file lokal (relatif ke media/soal_images/)
                        from django.conf import settings
                        full_path = os.path.join(settings.MEDIA_ROOT, 'soal_images', gambar_path)
                        if os.path.exists(full_path):
                            with open(full_path, 'rb') as img_file:
                                soal_image = File(img_file, name=os.path.basename(gambar_path))
                        else:
                            errors.append(f"Baris {row_num}: File gambar tidak ditemukan: {gambar_path}")
                
                # buat soal baru
                soal = Soal.objects.create(
                    subtest=subtest,
                    soal_text=soal_text or "",  # Bisa kosong jika hanya gambar
                    option_a=option_a,
                    option_b=option_b,
                    option_c=option_c,
                    option_d=option_d,
                    option_e=option_e,
                    correct_answer=kunci,
                )
                
                # Set gambar jika ada
                if soal_image:
                    soal.soal_image = soal_image
                    soal.save()
                
                created += 1
                
            except Exception as e:
                errors.append(f"Baris {row_num}: {str(e)}")
        
        wb.close()
        
        return Response(
            {
                'success': True,
                'created': created,
                'errors': errors[:10],  # batasi error yang ditampilkan
                'total_errors': len(errors),
            },
            status=status.HTTP_200_OK
        )
        
    except Exception as e:
        return Response(
            {'error': f'Gagal membaca file Excel: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
def submit_jawaban(request):
    """
    POST /api/submit-jawaban/
    
    Submit jawaban user untuk suatu subtest dan hitung skor.
    
    Body: {
        "username": "user123",
        "subtest_code": "LBI",
        "batch_id": "batch-1",
        "jawaban": {"1": "A", "2": "B", "3": "C", ...},  # {soal_id: jawaban}
        "durasi_detik": 1800  # optional
    }
    """
    username = request.data.get('username')
    subtest_code = request.data.get('subtest_code', '').upper()
    batch_id = request.data.get('batch_id', '')
    jawaban = request.data.get('jawaban', {})
    durasi_detik = request.data.get('durasi_detik')
    
    import logging
    logger = logging.getLogger(__name__)
    logger.info("=" * 50)
    logger.info(f"SUBMIT JAWABAN - Username: {username}, Subtest: {subtest_code}, Batch: {batch_id}")
    logger.info(f"Received jawaban type: {type(jawaban)}, count: {len(jawaban) if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban keys (first 10): {list(jawaban.keys())[:10] if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban sample: {dict(list(jawaban.items())[:5]) if isinstance(jawaban, dict) and jawaban else 'EMPTY'}")
    
    if not username or not subtest_code or not batch_id:
        return Response(
            {'error': 'username, subtest_code, dan batch_id wajib'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        subtest = Subtest.objects.get(code=subtest_code)
    except Subtest.DoesNotExist:
        return Response(
            {'error': f'Subtest dengan code {subtest_code} tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # ambil atau buat HasilTryout
    hasil, created = HasilTryout.objects.get_or_create(
        user=user,
        subtest=subtest,
        batch_id=batch_id,
        defaults={
            'jawaban': {},
            'waktu_mulai': timezone.now(),
        }
    )
    
    # update jawaban
    # jawaban dari frontend bisa dalam format {index: "A"} atau {soal_id: "A"}
    # kita perlu convert ke format {soal_id: "A"}
    jawaban_final = {}
    
    # ambil soal untuk subtest ini, filter by batch jika batch_id ada
    # Ini penting untuk memastikan jawaban disimpan sesuai dengan soal di batch yang dikerjakan
    try:
        if batch_id:
            # Coba filter by batch dulu
            batch_id_trimmed = str(batch_id).strip()
            soal_list = list(Soal.objects.filter(
                subtest=subtest,
                batch__batch_id=batch_id_trimmed
            ).select_related('batch').order_by('id'))
            logger.info(f"Filtering soal by batch_id='{batch_id_trimmed}', found {len(soal_list)} soal(s)")
            # Jika tidak ada soal dengan batch tersebut, ambil semua soal subtest
            if len(soal_list) == 0:
                logger.warning(f"No soal found for batch '{batch_id_trimmed}', using all soal for subtest {subtest.code}")
                soal_list = list(Soal.objects.filter(subtest=subtest).order_by('id'))
        else:
            # Tidak ada batch_id, ambil semua soal subtest
            soal_list = list(Soal.objects.filter(subtest=subtest).order_by('id'))
    except Exception as e:
        logger.error(f"Error filtering soal by batch: {e}", exc_info=True)
        # Fallback: ambil semua soal subtest
        soal_list = list(Soal.objects.filter(subtest=subtest).order_by('id'))
    
    # Buat mapping soal_id untuk validasi
    soal_id_set = {str(soal.id) for soal in soal_list}
    soal_id_list = [str(soal.id) for soal in soal_list]  # Untuk debugging
    
    logger.info(f"Processing jawaban for subtest {subtest_code}, batch {batch_id}")
    logger.info(f"Total soal: {len(soal_list)}")
    logger.info(f"Soal IDs in database (first 10): {soal_id_list[:10]}")
    logger.info(f"Received jawaban type: {type(jawaban)}, count: {len(jawaban) if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban keys (first 10): {list(jawaban.keys())[:10] if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban sample (first 5): {dict(list(jawaban.items())[:5]) if isinstance(jawaban, dict) else 'N/A'}")
    
    if not jawaban or (isinstance(jawaban, dict) and len(jawaban) == 0):
        logger.error("WARNING: Jawaban kosong atau tidak ada!")
        return Response(
            {'error': 'Jawaban kosong. Pastikan Anda sudah mengisi jawaban sebelum submit.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    matched_count = 0
    converted_count = 0
    skipped_count = 0
    
    # Debug: print semua soal_id yang ada di database
    logger.info(f"🔍 All soal_ids in database for {subtest_code}: {sorted([int(sid) for sid in soal_id_list])[:20]}")
    
    for key, value in jawaban.items():
        key_str = str(key).strip()
        value_str = str(value).strip().upper() if value else ""
        
        if not value_str or value_str not in ['A', 'B', 'C', 'D', 'E']:
            logger.warning(f"⚠️ Skipping invalid value for key {key_str}: '{value_str}' (must be A/B/C/D/E)")
            skipped_count += 1
            continue
        
        # Prioritas 1: Jika key_str ada di soal_id_set, langsung pakai sebagai soal_id
        if key_str in soal_id_set:
            jawaban_final[key_str] = value_str
            matched_count += 1
            logger.info(f"✅ Matched soal_id directly: {key_str} -> {value_str}")
            continue
        
        # Prioritas 2: Coba convert key sebagai integer (bisa jadi soal_id atau index)
        try:
            key_int = int(key_str)
            
            # Cek apakah key_int adalah soal_id yang valid
            if key_int in [int(sid) for sid in soal_id_list]:
                # Ini adalah soal_id yang valid, tapi dalam format integer
                jawaban_final[str(key_int)] = value_str
                matched_count += 1
                logger.info(f"✅ Matched soal_id (as int): {key_int} -> {value_str}")
                continue
            
            # Jika bukan soal_id, coba sebagai index (0-based)
            if 0 <= key_int < len(soal_list):
                # Ini adalah index, convert ke soal_id
                soal_id = str(soal_list[key_int].id)
                jawaban_final[soal_id] = value_str
                converted_count += 1
                logger.info(f"✅ Converted index {key_int} to soal_id {soal_id} -> {value_str}")
            else:
                # Key adalah angka tapi di luar range index dan bukan soal_id
                logger.warning(f"❌ Skipping key out of range: {key_str} (not a valid soal_id, max index: {len(soal_list)-1})")
                skipped_count += 1
        except (ValueError, TypeError):
            # Key bukan angka dan tidak ada di soal_id_set, skip
            logger.warning(f"❌ Skipping invalid key (not number, not in soal_id_set): {key_str}")
            skipped_count += 1
    
    logger.info(f"📊 Conversion summary: matched={matched_count}, converted={converted_count}, skipped={skipped_count}, final_count={len(jawaban_final)}")
    
    logger.info(f"Final jawaban_final: {len(jawaban_final)} entries")
    logger.info(f"Sample jawaban_final keys: {list(jawaban_final.keys())[:10]}")
    logger.info(f"Sample jawaban_final values: {list(jawaban_final.values())[:10]}")
    
    hasil.jawaban = jawaban_final
    logger.info(f"Jawaban assigned to hasil.jawaban, total: {len(hasil.jawaban)}")
    hasil.waktu_selesai = timezone.now()
    if durasi_detik:
        hasil.durasi_detik = durasi_detik
    
    # Simpan dulu sebelum hitung skor
    hasil.save(update_fields=['jawaban', 'waktu_selesai', 'durasi_detik'])
    logger.info(f"✅ HasilTryout saved with jawaban count: {len(hasil.jawaban)}")
    
    # hitung skor
    hasil.hitung_skor()
    logger.info(f"✅ Skor calculated: benar={hasil.jumlah_benar}, salah={hasil.jumlah_salah}, kosong={hasil.jumlah_kosong}, skor={hasil.skor}%")
    
    # Simpan hasil skor ke database
    hasil.save(update_fields=['jumlah_benar', 'jumlah_salah', 'jumlah_kosong', 'skor'])
    logger.info(f"✅ HasilTryout updated with skor calculation")
    
    # Pastikan skor dalam range 0-100
    skor_value = float(hasil.skor) if hasil.skor is not None else 0.0
    if skor_value < 0:
        skor_value = 0.0
    elif skor_value > 100:
        skor_value = 100.0
    
    return Response(
        {
            'success': True,
            'hasil': {
                'id': hasil.id,
                'subtest_code': subtest.code,
                'subtest_nama': subtest.nama_subtest,
                'batch_id': hasil.batch_id,
                'jumlah_benar': hasil.jumlah_benar,
                'jumlah_salah': hasil.jumlah_salah,
                'jumlah_kosong': hasil.jumlah_kosong,
                'skor': round(skor_value),  # Bulatkan ke bilangan bulat
                'waktu_selesai': hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None,
            }
        },
        status=status.HTTP_200_OK
    )


@api_view(['GET'])
def admin_dashboard(request):
    """
    GET /api/admin/dashboard/
    
    Mengembalikan statistik dan data untuk dashboard admin.
    Hanya bisa diakses oleh admin (is_staff=True).
    
    Query param: ?username=admin (wajib)
    """
    # Validasi admin dari username
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
        if not user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Statistik umum
    total_users = User.objects.count()
    total_students = User.objects.filter(is_staff=False).count()
    total_admins = User.objects.filter(is_staff=True).count()
    
    # Statistik subtest
    total_subtests = Subtest.objects.count()
    total_soal = Soal.objects.count()
    
    # Statistik hasil tryout
    total_hasil = HasilTryout.objects.count()
    hasil_dengan_skor = HasilTryout.objects.exclude(skor=0).count()
    avg_skor = HasilTryout.objects.aggregate(avg=Avg('skor'))['avg'] or 0
    
    # Statistik per subtest
    subtest_stats = []
    for subtest in Subtest.objects.all():
        hasil_subtest = HasilTryout.objects.filter(subtest=subtest)
        subtest_stats.append({
            'code': subtest.code,
            'nama': subtest.nama_subtest,
            'jumlah_soal': subtest.jumlah_soal,
            'total_pengerjaan': hasil_subtest.count(),
            'avg_skor': round(float(hasil_subtest.aggregate(avg=Avg('skor'))['avg'] or 0)),
        })
    
    # Top 10 users dengan skor tertinggi
    top_users = HasilTryout.objects.values(
        'user__username', 'user__first_name', 'user__last_name'
    ).annotate(
        total_skor=Avg('skor'),
        total_pengerjaan=Count('id')
    ).order_by('-total_skor')[:10]
    
    return Response({
        'stats': {
            'users': {
                'total': total_users,
                'students': total_students,
                'admins': total_admins,
            },
            'subtests': {
                'total': total_subtests,
                'total_soal': total_soal,
            },
            'hasil_tryout': {
                'total': total_hasil,
                'dengan_skor': hasil_dengan_skor,
                'avg_skor': round(float(avg_skor)),
            },
        },
        'subtest_stats': subtest_stats,
        'top_users': [
            {
                'username': u['user__username'],
                'name': f"{u['user__first_name']} {u['user__last_name']}".strip() or u['user__username'],
                'avg_skor': round(float(u['total_skor'])),
                'total_pengerjaan': u['total_pengerjaan'],
            }
            for u in top_users
        ],
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def admin_list_soal(request):
    """
    GET /api/admin/soal/
    
    List semua soal dengan pagination.
    Query params: ?username=admin&page=1&limit=50&subtest_code=PPU&search=...
    """
    # Validasi admin
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
        if not user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from django.core.paginator import Paginator
    
    subtest_code = request.GET.get('subtest_code', '').upper()
    search = request.GET.get('search', '').strip()
    batch_id = request.GET.get('batch_id', '').strip()
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 50))
    
    queryset = Soal.objects.select_related('subtest', 'batch').all()
    
    if batch_id:
        try:
            batch_id_int = int(batch_id)
            queryset = queryset.filter(batch__id=batch_id_int)
        except ValueError:
            pass  # Invalid batch_id, ignore
    
    if subtest_code:
        queryset = queryset.filter(subtest__code=subtest_code)
    
    if search:
        queryset = queryset.filter(
            Q(soal_text__icontains=search) |
            Q(option_a__icontains=search) |
            Q(option_b__icontains=search) |
            Q(option_c__icontains=search) |
            Q(option_d__icontains=search) |
            Q(option_e__icontains=search)
        )
    
    paginator = Paginator(queryset, limit)
    page_obj = paginator.get_page(page)
    
    data = []
    for soal in page_obj:
        image_url = None
        if soal.soal_image:
            request_scheme = request.scheme if hasattr(request, 'scheme') else 'http'
            request_host = request.get_host() if hasattr(request, 'get_host') else 'localhost:8000'
            # Pastikan URL path di-encode dengan benar
            media_path = str(soal.soal_image.url) if hasattr(soal.soal_image, 'url') else ''
            # URL dari Django sudah dalam format yang benar, langsung gabungkan
            image_url = f"{request_scheme}://{request_host}{media_path}"
        
        data.append({
            'id': soal.id,
            'subtest_code': soal.subtest.code,
            'subtest_nama': soal.subtest.nama_subtest,
            'soal_text': soal.soal_text[:200] + '...' if len(soal.soal_text) > 200 else soal.soal_text,
            'soal_image': image_url,
            'has_image': bool(soal.soal_image),
            'option_a': soal.option_a,
            'option_b': soal.option_b,
            'option_c': soal.option_c,
            'option_d': soal.option_d,
            'option_e': soal.option_e,
            'correct_answer': soal.correct_answer,
            'created_at': soal.created_at.isoformat(),
        })
    
    return Response({
        'results': data,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_prev': page_obj.has_previous(),
        },
    }, status=status.HTTP_200_OK)




@csrf_exempt
@require_http_methods(['POST'])
def admin_create_soal(request):
    """
    POST /api/admin/soal/create/

    Buat soal baru via API (untuk digunakan oleh frontend admin SPA).
    Query param: ?username=admin (wajib)
    """
    import logging
    import json
    from django.http import JsonResponse
    
    logger = logging.getLogger(__name__)
    logger.info("=== admin_create_soal called ===")
    
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return JsonResponse(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=400
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            logger.warning(f"Access denied: user {username} is not staff")
            return JsonResponse(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=403
            )
    except User.DoesNotExist:
        logger.warning(f"Admin user not found: {username}")
        return JsonResponse(
            {'error': 'Admin user tidak ditemukan'},
            status=404
        )
    
    logger.info(f"Admin user '{username}' allowed to create soal")

    try:
        # Support both JSON and multipart/form-data (with files)
        data = {}
        if request.content_type and request.content_type.startswith('multipart'):
            # multipart: fields in request.POST, files in request.FILES
            data = request.POST.dict()
        else:
            if request.body:
                data = json.loads(request.body)
            else:
                data = {}
    except (json.JSONDecodeError, ValueError) as e:
        logger.error(f"Failed to parse request body: {e}")
        return JsonResponse({'error': 'Invalid request body'}, status=400)

    pertanyaan = data.get('pertanyaan') or data.get('soal_text') or ''
    subtest_code = (data.get('subtest') or data.get('subtest_code') or '').upper()
    pilihan = data.get('pilihan') or ''
    kunci = (data.get('kunci_jawaban') or data.get('correct_answer') or '').upper()
    batch_id = data.get('batch_id')  # ID batch (int) atau None

    if not pertanyaan or not subtest_code:
        logger.warning(f"Missing fields: pertanyaan='{pertanyaan}', subtest_code='{subtest_code}'")
        return JsonResponse({'error': 'Field pertanyaan dan subtest wajib.'}, status=400)

    try:
        subtest = Subtest.objects.get(code=subtest_code)
    except Subtest.DoesNotExist:
        logger.error(f"Subtest not found: {subtest_code}")
        return JsonResponse({'error': f'Subtest dengan kode {subtest_code} tidak ditemukan.'}, status=404)

    # Parse pilihan jika diberikan sebagai pipe-separated string
    options = ['', '', '', '', '']
    if pilihan:
        parts = [p.strip() for p in str(pilihan).split('|')]
        for i in range(min(5, len(parts))):
            # remove leading 'A. ' if present
            part = parts[i]
            if len(part) > 2 and part[1] == '.' and part[0].isalpha():
                part = part[2:].strip()
            options[i] = part
    else:
        # coba ambil option_a..e dari payload
        for idx, field in enumerate(['option_a', 'option_b', 'option_c', 'option_d', 'option_e']):
            options[idx] = data.get(field, '')

    # fallback jika beberapa pilihan kosong
    for i in range(5):
        if options[i] is None:
            options[i] = ''

    if not kunci or kunci not in ['A', 'B', 'C', 'D', 'E']:
        kunci = 'A'

    # Handle batch assignment
    batch = None
    if batch_id:
        try:
            batch_id_int = int(batch_id)
            batch = Batch.objects.get(id=batch_id_int)
            logger.info(f"Assigning soal to batch: {batch.batch_id}")
        except (ValueError, Batch.DoesNotExist):
            logger.warning(f"Batch ID {batch_id} tidak valid, membuat soal tanpa batch")
    
    # Create Soal instance first (so we can assign soal_image using model field)
    soal = Soal.objects.create(
        subtest=subtest,
        batch=batch,
        soal_text=pertanyaan,
        option_a=options[0] or '-',
        option_b=options[1] or '-',
        option_c=options[2] or '-',
        option_d=options[3] or '-',
        option_e=options[4] or '-',
        correct_answer=kunci,
    )

    # Handle uploaded files (soal image + option images)
    try:
        from django.core.files.storage import default_storage
        from django.conf import settings
        # soal image
        soal_img = request.FILES.get('soal_image') or request.FILES.get('soal_img')
        if soal_img:
            soal.soal_image.save(soal_img.name, soal_img, save=True)

        # option images: option_a_image, option_b_image, ...
        for idx, opt_field in enumerate(['option_a', 'option_b', 'option_c', 'option_d', 'option_e']):
            file_key = f"option_{opt_field[-1]}_image"  # option_a -> a
            f = request.FILES.get(file_key)
            if f:
                # save file to storage and set option text to URL
                path = default_storage.save(f"option_images/{f.name}", f)
                try:
                    url = default_storage.url(path)
                except Exception:
                    url = (settings.MEDIA_URL.rstrip('/') + '/') + path
                setattr(soal, opt_field, url)
        soal.save()
    except Exception as e:
        logger.exception(f"Failed to save uploaded files: {e}")
    
    logger.info(f"Soal created successfully: id={soal.id}, subtest={subtest.code}")

    return JsonResponse({
        'success': True,
        'soal': {
            'id': soal.id,
            'subtest_code': subtest.code,
            'soal_text': soal.soal_text,
            'correct_answer': soal.correct_answer,
        }
    }, status=201)


@csrf_exempt
@require_http_methods(['PUT', 'PATCH'])
def admin_update_soal(request, soal_id):
    """
    PUT/PATCH /api/admin/soal/<soal_id>/update/
    Update soal fields. Auth: Django session (is_staff required).
    Query param: ?username=admin (wajib)
    """
    import logging
    import json

    logger = logging.getLogger(__name__)
    logger.info(f"=== admin_update_soal called for id={soal_id} ===")
    
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return JsonResponse(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=400
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            logger.warning(f"Access denied for update_soal: user {username} is not staff")
            return JsonResponse(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=403
            )
    except User.DoesNotExist:
        logger.warning(f"Admin user not found: {username}")
        return JsonResponse(
            {'error': 'Admin user tidak ditemukan'},
            status=404
        )
    
    logger.info(f"Admin user {username} allowed to update soal")

    try:
        soal = Soal.objects.get(id=soal_id)
    except Soal.DoesNotExist:
        return JsonResponse({'error': 'Soal tidak ditemukan.'}, status=404)

    try:
        data = {}
        if request.content_type and request.content_type.startswith('multipart'):
            data = request.POST.dict()
        else:
            data = json.loads(request.body) if request.body else {}
    except (ValueError, json.JSONDecodeError) as e:
        logger.error(f"JSON parse error: {e}")
        return JsonResponse({'error': 'Invalid JSON body'}, status=400)

    # Update fields if provided
    pertanyaan = data.get('pertanyaan') or data.get('soal_text')
    subtest_code = (data.get('subtest') or data.get('subtest_code') or '')
    pilihan = data.get('pilihan')
    kunci = (data.get('kunci_jawaban') or data.get('correct_answer') or '').upper()

    if pertanyaan is not None:
        soal.soal_text = pertanyaan

    if subtest_code:
        try:
            subtest = Subtest.objects.get(code=subtest_code.upper())
            soal.subtest = subtest
        except Subtest.DoesNotExist:
            return JsonResponse({'error': f'Subtest {subtest_code} tidak ditemukan.'}, status=404)

    if pilihan is not None:
        # parse pilihan pipe-separated or accept option_a..e
        options = ['', '', '', '', '']
        if pilihan:
            parts = [p.strip() for p in str(pilihan).split('|')]
            for i in range(min(5, len(parts))):
                part = parts[i]
                if len(part) > 2 and part[1] == '.' and part[0].isalpha():
                    part = part[2:].strip()
                options[i] = part
        else:
            for idx, field in enumerate(['option_a', 'option_b', 'option_c', 'option_d', 'option_e']):
                options[idx] = data.get(field, '')

        soal.option_a = options[0] or '-'
        soal.option_b = options[1] or '-'
        soal.option_c = options[2] or '-'
        soal.option_d = options[3] or '-'
        soal.option_e = options[4] or '-'

    # Handle file uploads for update
    try:
        from django.core.files.storage import default_storage
        from django.conf import settings
        soal_img = request.FILES.get('soal_image') or request.FILES.get('soal_img')
        if soal_img:
            # replace existing soal_image
            soal.soal_image.save(soal_img.name, soal_img, save=False)

        for idx, opt_field in enumerate(['option_a', 'option_b', 'option_c', 'option_d', 'option_e']):
            file_key = f"option_{opt_field[-1]}_image"
            f = request.FILES.get(file_key)
            if f:
                path = default_storage.save(f"option_images/{f.name}", f)
                try:
                    url = default_storage.url(path)
                except Exception:
                    url = (settings.MEDIA_URL.rstrip('/') + '/') + path
                setattr(soal, opt_field, url)
    except Exception as e:
        logger.exception(f"Failed to process uploaded files for update: {e}")

    if kunci:
        if kunci not in ['A', 'B', 'C', 'D', 'E']:
            kunci = 'A'
        soal.correct_answer = kunci

    soal.save()

    logger.info(f"Soal updated: id={soal.id}")
    return JsonResponse({'success': True, 'soal': {
        'id': soal.id,
        'subtest_code': soal.subtest.code,
        'soal_text': soal.soal_text,
        'correct_answer': soal.correct_answer,
    }}, status=200)


@csrf_exempt
@require_http_methods(['DELETE'])
def admin_delete_soal(request, soal_id):
    """
    DELETE /api/admin/soal/<soal_id>/delete/
    Delete soal. Auth: Django session (is_staff required).
    Query param: ?username=admin (wajib)
    """
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"=== admin_delete_soal called for id={soal_id} ===")

    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return JsonResponse(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=400
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            logger.warning(f"Access denied for delete_soal: user {username} is not staff")
            return JsonResponse(
                {'error': 'Akses ditolak. Hanya admin.'},
                status=403
            )
    except User.DoesNotExist:
        logger.warning(f"Admin user not found: {username}")
        return JsonResponse(
            {'error': 'Admin user tidak ditemukan'},
            status=404
        )
    
    logger.info(f"Admin user {username} allowed to delete soal")

    try:
        soal = Soal.objects.get(id=soal_id)
    except Soal.DoesNotExist:
        return JsonResponse({'error': 'Soal tidak ditemukan.'}, status=404)

    soal.delete()
    logger.info(f"Soal deleted: id={soal_id}")
    return JsonResponse({'success': True}, status=200)
@api_view(['GET'])
def admin_list_users(request):
    """
    GET /api/admin/users/
    
    List semua users dengan statistik.
    Query param: ?username=admin (wajib)
    """
    # Validasi admin
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
        if not user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    users = User.objects.annotate(
        total_hasil=Count('hasil_tryout'),
        avg_skor=Avg('hasil_tryout__skor')
    ).order_by('-date_joined')

    from django.core.cache import cache
    data = []
    for user in users:
        # Retrieve password from cache if available (newly created users)
        cached_password = cache.get(f'user_password_{user.id}')
        data.append({
            'id': user.id,
            'username': user.username,
            'password': cached_password or '-',
            'name': user.get_full_name() or user.username,
            'email': user.email or '',
            'is_staff': user.is_staff,
            'is_active': user.is_active,
            'date_joined': user.date_joined.isoformat(),
        })
    
    return Response(data, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
def admin_create_user(request):
    """
    POST /api/admin/users/create/
    Create a new user. Auth: session and staff required.
    Query param: ?username=admin (wajib)
    Expected JSON: { username, password, name?, email?, is_staff?: bool }
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )

    data = request.data
    username = (data.get('username') or '').strip()
    password = data.get('password')
    name = data.get('name') or ''
    email = data.get('email') or ''
    is_staff = bool(data.get('is_staff', False))

    if not username or not password:
        return Response({'error': 'username dan password wajib.'}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.filter(username=username).exists():
        return Response({'error': 'Username sudah digunakan.'}, status=status.HTTP_400_BAD_REQUEST)

    user = User.objects.create_user(username=username, password=password, email=email)
    if name:
        # try to split name into first/last
        parts = name.split(None, 1)
        user.first_name = parts[0]
        if len(parts) > 1:
            user.last_name = parts[1]
    user.is_staff = is_staff
    user.is_active = True
    user.save()

    # Store password in cache for display (expires in 10 years - effectively permanent for admin viewing)
    from django.core.cache import cache
    cache.set(f'user_password_{user.id}', password, timeout=315360000)  # 10 years

    return Response({'success': True, 'user': {
        'id': user.id,
        'username': user.username,
        'password': password,
        'name': user.get_full_name() or user.username,
        'email': user.email,
        'is_staff': user.is_staff,
        'is_active': user.is_active,
    }}, status=status.HTTP_201_CREATED)


@csrf_exempt
@api_view(['DELETE'])
def admin_delete_user(request, user_id):
    """
    DELETE /api/admin/users/<user_id>/delete/
    Delete a user. Admin only.
    Query param: ?username=admin (wajib)
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )

    if admin_user.id == user_id:
        return Response({'error': 'Tidak bisa menghapus akun sendiri.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        u = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

    u.delete()
    return Response({'success': True}, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
def admin_toggle_user_active(request, user_id):
    """
    POST /api/admin/users/<user_id>/toggle-active/
    Toggle user's is_active state (or set using payload {is_active: true/false}).
    Query param: ?username=admin (wajib)
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )

    try:
        u = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

    if admin_user.id == user_id:
        return Response({'error': 'Tidak bisa mengubah status akun sendiri.'}, status=status.HTTP_400_BAD_REQUEST)

    # payload optional
    payload_is_active = request.data.get('is_active', None)
    if payload_is_active is None:
        u.is_active = not u.is_active
    else:
        u.is_active = bool(payload_is_active)

    u.save()
    return Response({'success': True, 'user': {'id': u.id, 'is_active': u.is_active}}, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
def admin_reveal_password(request, user_id):
    """
    POST /api/admin/users/<user_id>/reveal-password/
    Generate a new password for user and store it in cache for admin viewing.
    Admin only.
    Query param: ?username=admin (wajib)
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )

    try:
        u = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

    # Generate a random password (8-12 characters, alphanumeric)
    import secrets
    import string
    alphabet = string.ascii_letters + string.digits
    new_password = ''.join(secrets.choice(alphabet) for i in range(10))
    
    # Set the new password for the user
    u.set_password(new_password)
    u.save()
    
    # Store password in cache for display (expires in 10 years)
    from django.core.cache import cache
    cache.set(f'user_password_{u.id}', new_password, timeout=315360000)  # 10 years
    
    return Response({
        'success': True,
        'password': new_password,
        'user': {
            'id': u.id,
            'username': u.username,
        }
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
def admin_update_user_password(request, user_id):
    """
    POST /api/admin/users/<user_id>/update-password/
    Update user's password. Admin only.
    Query param: ?username=admin (wajib)
    Expected JSON: { "password": "new_password" }
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )

    try:
        u = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

    password = request.data.get('password', '').strip()
    if not password:
        return Response({'error': 'Password wajib.'}, status=status.HTTP_400_BAD_REQUEST)

    # Set the new password
    u.set_password(password)
    u.save()
    
    # Store password in cache for display (expires in 10 years)
    from django.core.cache import cache
    cache.set(f'user_password_{u.id}', password, timeout=315360000)  # 10 years
    
    return Response({
        'success': True,
        'password': password,
        'user': {
            'id': u.id,
            'username': u.username,
        }
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
def admin_import_users_excel(request):
    """
    POST /api/admin/users/import-excel/
    
    Import users dari file Excel.
    Format: Username | Password | Nama
    
    Body: multipart/form-data dengan field 'file' (file Excel .xlsx)
    Query param: ?username=admin (wajib)
    """
    # Validasi admin via query parameter
    admin_username = request.GET.get('username', '').strip()
    if not admin_username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=admin_username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if load_workbook is None:
        return Response(
            {'error': 'openpyxl belum terinstall. Install dengan: pip install openpyxl'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    if 'file' not in request.FILES:
        return Response(
            {'error': 'File Excel tidak ditemukan. Kirim dengan field "file".'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    excel_file = request.FILES['file']
    
    # Validasi ekstensi
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'File harus berformat .xlsx atau .xls'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Baca file Excel dengan openpyxl
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        
        # Baca header
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        expected_headers = ["Username", "Password", "Nama"]
        
        # Validasi header (case-insensitive)
        header_lower = [str(h).strip().lower() if h else "" for h in header_row[:3]]
        expected_lower = [h.lower() for h in expected_headers]
        
        if header_lower != expected_lower:
            return Response(
                {
                    'error': f'Header tidak sesuai. Diharapkan: {", ".join(expected_headers)}',
                    'found': header_lower
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created = 0
        skipped = 0
        errors = []
        created_users = []
        
        # Import password ke cache
        from django.core.cache import cache
        
        # Proses setiap baris data
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip baris kosong
            if not any(row[:3]):
                continue
            
            try:
                username = str(row[0]).strip() if row[0] else ""
                password = str(row[1]).strip() if row[1] else ""
                name = str(row[2]).strip() if row[2] else ""
                
                # Validasi
                if not username:
                    errors.append(f"Baris {row_num}: Username kosong")
                    continue
                
                if not password:
                    errors.append(f"Baris {row_num}: Password kosong")
                    continue
                
                # Cek apakah username sudah ada
                if User.objects.filter(username=username).exists():
                    skipped += 1
                    errors.append(f"Baris {row_num}: Username '{username}' sudah digunakan (dilewati)")
                    continue
                
                # Buat user baru
                user = User.objects.create_user(username=username, password=password)
                
                # Set nama
                if name:
                    parts = name.split(None, 1)
                    user.first_name = parts[0]
                    if len(parts) > 1:
                        user.last_name = parts[1]
                
                user.is_active = True
                user.is_staff = False  # Default: bukan admin
                user.save()
                
                # Store password in cache for display
                cache.set(f'user_password_{user.id}', password, timeout=315360000)  # 10 years
                
                created += 1
                created_users.append({
                    'id': user.id,
                    'username': user.username,
                    'name': user.get_full_name() or user.username,
                    'password': password,
                })
                
            except Exception as e:
                errors.append(f"Baris {row_num}: {str(e)}")
                continue
        
        return Response({
            'success': True,
            'created': created,
            'skipped': skipped,
            'errors': errors,
            'users': created_users,
            'message': f'Berhasil mengimport {created} user. {skipped} user dilewati (username sudah ada).'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Error membaca file Excel: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def admin_list_hasil(request):
    """
    GET /api/admin/hasil/
    
    List semua hasil tryout dengan filter.
    Query params: ?username=admin&search_username=...&subtest_code=...&page=1&limit=50
    
    - username: Username admin untuk validasi (required)
    - search_username: Username user untuk filter hasil tryout (optional)
    - subtest_code: Kode subtest untuk filter (optional)
    - page: Nomor halaman (default: 1)
    - limit: Jumlah item per halaman (default: 50)
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from django.core.paginator import Paginator
    
    # Parameter untuk filter user (bukan admin)
    search_username = request.GET.get('search_username', '').strip()
    subtest_code = request.GET.get('subtest_code', '').upper()
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 50))
    
    queryset = HasilTryout.objects.select_related('user', 'subtest').all()
    
    # Filter berdasarkan username user yang dicari
    if search_username:
        queryset = queryset.filter(user__username__icontains=search_username)
    
    if subtest_code:
        queryset = queryset.filter(subtest__code=subtest_code)
    
    queryset = queryset.order_by('-waktu_selesai', '-created_at')
    
    paginator = Paginator(queryset, limit)
    page_obj = paginator.get_page(page)
    
    data = []
    for hasil in page_obj:
        data.append({
            'id': hasil.id,
            'username': hasil.user.username,
            'user_name': hasil.user.get_full_name() or hasil.user.username,
            'subtest_code': hasil.subtest.code,
            'subtest_nama': hasil.subtest.nama_subtest,
            'batch_id': hasil.batch_id,
            'jumlah_benar': hasil.jumlah_benar,
            'jumlah_salah': hasil.jumlah_salah,
            'jumlah_kosong': hasil.jumlah_kosong,
            'skor': round(float(hasil.skor)) if hasil.skor else 0,
            'waktu_selesai': hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None,
            'durasi_detik': hasil.durasi_detik,
        })
    
    return Response({
        'results': data,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_prev': page_obj.has_previous(),
        },
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def admin_detail_jawaban(request, hasil_id):
    """
    GET /api/admin/hasil/<hasil_id>/detail/
    
    Ambil detail jawaban untuk hasil tryout tertentu.
    Query params: ?username=admin
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        # Validasi admin via query parameter (konsisten dengan endpoint lain)
        username = request.GET.get('username', '').strip()
        if not username:
            return Response(
                {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            admin_user = User.objects.get(username=username)
            if not admin_user.is_staff:
                return Response(
                    {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except User.DoesNotExist:
            return Response(
                {'error': 'Admin user tidak ditemukan'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            hasil = HasilTryout.objects.select_related('user', 'subtest').get(id=hasil_id)
        except HasilTryout.DoesNotExist:
            return Response(
                {'error': 'Hasil tryout tidak ditemukan'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Ambil semua soal untuk subtest ini
        # Filter by batch jika batch_id ada, tapi jika tidak ada soal dengan batch tersebut, ambil semua soal subtest
        logger.info(f"Fetching soal for hasil_id={hasil_id}, subtest={hasil.subtest.code}, batch_id='{hasil.batch_id}'")
        try:
            batch_id_trimmed = str(hasil.batch_id).strip() if hasil.batch_id else ''
            if batch_id_trimmed:
                # Coba filter by batch dulu - gunakan select_related untuk optimasi
                logger.info(f"Filtering soal by batch_id='{batch_id_trimmed}'")
                try:
                    soal_list = Soal.objects.filter(
                        subtest=hasil.subtest,
                        batch__batch_id=batch_id_trimmed
                    ).select_related('subtest', 'batch').order_by('id')
                    soal_count = soal_list.count()
                    logger.info(f"Found {soal_count} soal(s) for batch '{batch_id_trimmed}'")
                    # Jika tidak ada soal dengan batch tersebut, ambil semua soal subtest
                    if soal_count == 0:
                        logger.warning(f"No soal found for batch '{batch_id_trimmed}', using all soal for subtest {hasil.subtest.code}")
                        soal_list = Soal.objects.filter(subtest=hasil.subtest).select_related('subtest', 'batch').order_by('id')
                        logger.info(f"Fallback: Found {soal_list.count()} soal(s) for subtest {hasil.subtest.code}")
                except Exception as filter_error:
                    logger.error(f"Error filtering by batch_id '{batch_id_trimmed}': {filter_error}", exc_info=True)
                    # Fallback ke semua soal subtest
                    soal_list = Soal.objects.filter(subtest=hasil.subtest).select_related('subtest', 'batch').order_by('id')
                    logger.info(f"Fallback after filter error: Found {soal_list.count()} soal(s) for subtest {hasil.subtest.code}")
            else:
                logger.info(f"No batch_id provided, using all soal for subtest {hasil.subtest.code}")
                soal_list = Soal.objects.filter(subtest=hasil.subtest).select_related('subtest', 'batch').order_by('id')
                logger.info(f"Found {soal_list.count()} soal(s) for subtest {hasil.subtest.code}")
        except Exception as e:
            # Fallback: ambil semua soal subtest jika ada error
            logger.error(f"Error filtering soal by batch: {e}", exc_info=True)
            soal_list = Soal.objects.filter(subtest=hasil.subtest).select_related('subtest', 'batch').order_by('id')
            logger.info(f"Fallback after error: Found {soal_list.count()} soal(s) for subtest {hasil.subtest.code}")
        
        # Siapkan data detail jawaban
        detail_jawaban = []
        request_scheme = request.scheme if hasattr(request, 'scheme') else 'http'
        request_host = request.get_host() if hasattr(request, 'get_host') else 'localhost:8000'
        
        soal_list = list(soal_list)  # Convert to list to avoid multiple DB queries
        total_soal = len(soal_list)
        logger.info(f"Processing {total_soal} soal(s) for detail jawaban")
        logger.info(f"Jawaban dict keys count: {len(hasil.jawaban) if hasil.jawaban and isinstance(hasil.jawaban, dict) else 0}")
        
        if total_soal == 0:
            logger.warning(f"No soal found for subtest {hasil.subtest.code}, batch_id='{hasil.batch_id}'")
            return Response({
                'hasil_id': hasil.id,
                'username': hasil.user.username,
                'user_name': hasil.user.get_full_name() or hasil.user.username,
                'subtest_code': hasil.subtest.code,
                'subtest_nama': hasil.subtest.nama_subtest,
                'batch_id': hasil.batch_id,
                'skor': round(float(hasil.skor)) if hasil.skor else 0,
                'jumlah_benar': hasil.jumlah_benar or 0,
                'jumlah_salah': hasil.jumlah_salah or 0,
                'jumlah_kosong': hasil.jumlah_kosong or 0,
                'waktu_selesai': hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None,
                'durasi_detik': hasil.durasi_detik,
                'detail_jawaban': [],
                'warning': f'Tidak ada soal ditemukan untuk subtest {hasil.subtest.code}'
            }, status=status.HTTP_200_OK)
        
        for idx, soal in enumerate(soal_list, start=1):
            try:
                soal_id_str = str(soal.id)
                # Handle jawaban user dengan aman
                jawaban_user = ""
                if hasil.jawaban and isinstance(hasil.jawaban, dict):
                    jawaban_user = hasil.jawaban.get(soal_id_str, "")
                    if jawaban_user:
                        jawaban_user = str(jawaban_user).strip().upper()
                
                # Handle jawaban benar dengan aman
                jawaban_benar = ""
                if soal.correct_answer:
                    jawaban_benar = str(soal.correct_answer).strip().upper()
                
                # Tentukan status_jawaban (gunakan nama berbeda untuk menghindari konflik dengan status HTTP)
                if not jawaban_user:
                    status_jawaban = 'kosong'
                    is_benar = False
                elif jawaban_user == jawaban_benar:
                    status_jawaban = 'benar'
                    is_benar = True
                else:
                    status_jawaban = 'salah'
                    is_benar = False
                
                # URL gambar jika ada
                soal_image_url = None
                try:
                    if soal.soal_image and hasattr(soal.soal_image, 'url'):
                        image_url = soal.soal_image.url
                        if image_url:
                            soal_image_url = f"{request_scheme}://{request_host}{image_url}"
                except (ValueError, AttributeError) as e:
                    # ValueError occurs when image field is empty
                    # AttributeError occurs when url attribute doesn't exist
                    logger.debug(f"No image URL for soal {soal.id}: {e}")
                    soal_image_url = None
                except Exception as e:
                    logger.warning(f"Error getting soal image URL for soal {soal.id}: {e}")
                    soal_image_url = None
                
                # URL gambar untuk setiap opsi
                # Catatan: Gambar opsi disimpan sebagai URL di field option_a, option_b, dll (bukan field terpisah)
                option_images = {}
                
                # Helper function untuk cek apakah string adalah URL gambar
                def looks_like_image_url(text):
                    if not text:
                        return False
                    text = str(text).strip()
                    # Cek jika mengandung ekstensi gambar atau path media
                    return (
                        text.startswith('/media/') or
                        text.startswith('http://') or
                        text.startswith('https://') or
                        any(text.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg'])
                    )
                
                for opt_key in ['A', 'B', 'C', 'D', 'E']:
                    opt_field = f'option_{opt_key.lower()}'
                    try:
                        opt_value = getattr(soal, opt_field, '') or ''
                        if looks_like_image_url(opt_value):
                            # Jika URL relatif, tambahkan origin
                            if opt_value.startswith('/media/'):
                                option_images[opt_key] = f"{request_scheme}://{request_host}{opt_value}"
                            else:
                                option_images[opt_key] = opt_value
                    except Exception as e:
                        logger.warning(f"Error processing option {opt_key} for soal {soal.id}: {e}")
                
                # Siapkan pilihan: jika opsi adalah gambar, set teks ke None, jika teks gunakan teks
                pilihan = {}
                for opt_key in ['A', 'B', 'C', 'D', 'E']:
                    opt_field = f'option_{opt_key.lower()}'
                    try:
                        opt_value = getattr(soal, opt_field, '') or ''
                        # Jika opsi adalah gambar, set teks ke None (akan ditampilkan sebagai gambar)
                        if opt_key in option_images:
                            pilihan[opt_key] = None  # Akan ditampilkan sebagai gambar
                        else:
                            pilihan[opt_key] = str(opt_value)  # Teks biasa
                    except Exception as e:
                        logger.warning(f"Error getting option {opt_key} for soal {soal.id}: {e}")
                        pilihan[opt_key] = ''
                
                detail_jawaban.append({
                    'nomor': idx,
                    'soal_id': soal.id,
                    'pertanyaan': soal.soal_text or '',
                    'soal_image': soal_image_url,
                    'pilihan': pilihan,
                    'option_images': option_images,
                    'jawaban_user': jawaban_user if jawaban_user else None,
                    'jawaban_benar': jawaban_benar,
                    'status': status_jawaban,
                    'is_benar': is_benar,
                })
            except Exception as e:
                logger.error(f"Error processing soal {soal.id if soal else 'unknown'}: {e}", exc_info=True)
                # Skip soal ini dan lanjutkan ke soal berikutnya
                continue
        
        # Siapkan response data dengan error handling
        try:
            waktu_selesai_iso = hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None
        except Exception as e:
            logger.warning(f"Error formatting waktu_selesai: {e}")
            waktu_selesai_iso = None
        
        try:
            # Safely get user fields
            try:
                username = hasil.user.username if hasil.user else 'unknown'
            except Exception as e:
                logger.error(f"Error accessing user.username: {e}")
                username = 'unknown'
            
            try:
                user_name = (hasil.user.get_full_name() or hasil.user.username) if hasil.user else 'unknown'
            except Exception as e:
                logger.error(f"Error accessing user name: {e}")
                user_name = username
            
            # Safely get subtest fields
            try:
                subtest_code = hasil.subtest.code if hasil.subtest else ''
            except Exception as e:
                logger.error(f"Error accessing subtest.code: {e}")
                subtest_code = ''
            
            try:
                subtest_nama = hasil.subtest.nama_subtest if hasil.subtest else ''
            except Exception as e:
                logger.error(f"Error accessing subtest.nama_subtest: {e}")
                subtest_nama = ''
            
            response_data = {
                'hasil_id': hasil.id,
                'username': username,
                'user_name': user_name,
                'subtest_code': subtest_code,
                'subtest_nama': subtest_nama,
                'batch_id': hasil.batch_id or '',
                'skor': round(float(hasil.skor), 2) if hasil.skor is not None else 0.0,
                'jumlah_benar': hasil.jumlah_benar or 0,
                'jumlah_salah': hasil.jumlah_salah or 0,
                'jumlah_kosong': hasil.jumlah_kosong or 0,
                'waktu_selesai': waktu_selesai_iso,
                'durasi_detik': hasil.durasi_detik or 0,
                'detail_jawaban': detail_jawaban,
            }
            logger.info(f"Successfully prepared response with {len(detail_jawaban)} detail jawaban items")
            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error preparing response data: {e}", exc_info=True)
            raise
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        logger.error(f"Error in admin_detail_jawaban for hasil_id={hasil_id}: {e}")
        logger.error(f"Full traceback:\n{error_traceback}")
        return Response(
            {
                'error': f'Terjadi kesalahan saat mengambil detail jawaban: {str(e)}',
                'error_type': type(e).__name__,
                'detail': str(e)
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def admin_list_batches(request):
    """
    GET /api/admin/batches/
    List semua batches untuk admin.
    Query param: ?username=admin (wajib)
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    batches = Batch.objects.all().order_by('-date', '-created_at')
    data = []
    for batch in batches:
        data.append({
            'id': batch.id,
            'batch_id': batch.batch_id,
            'title': batch.title,
            'date': batch.date.isoformat(),
            'date_display': batch.date.strftime('%d %B %Y'),
            'deadline': batch.deadline.isoformat(),
            'deadline_display': batch.deadline.strftime('%d %B %Y'),
            'status': batch.status,
            'description': batch.description or '',
            'is_visible': batch.is_visible,
            'created_at': batch.created_at.isoformat(),
            'updated_at': batch.updated_at.isoformat(),
        })
    
    return Response({'results': data}, status=status.HTTP_200_OK)


@api_view(['POST'])
def admin_create_batch(request):
    """
    POST /api/admin/batches/create/
    Buat batch baru.
    Query param: ?username=admin (wajib)
    Body: { batch_id, title, date, deadline, status, description, is_visible }
    """
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    data = request.data
    batch_id = data.get('batch_id', '').strip()
    title = data.get('title', '').strip()
    date_str = data.get('date', '')
    deadline_str = data.get('deadline', '')
    status_val = data.get('status', 'locked')
    description = data.get('description', '').strip()
    is_visible = data.get('is_visible', True)
    
    if not batch_id or not title or not date_str or not deadline_str:
        return Response({'error': 'batch_id, title, date, dan deadline wajib.'}, status=status.HTTP_400_BAD_REQUEST)
    
    if Batch.objects.filter(batch_id=batch_id).exists():
        return Response({'error': 'Batch dengan ID tersebut sudah ada.'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        from datetime import datetime
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        deadline = datetime.strptime(deadline_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Format tanggal tidak valid. Gunakan YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
    
    batch = Batch.objects.create(
        batch_id=batch_id,
        title=title,
        date=date,
        deadline=deadline,
        status=status_val,
        description=description,
        is_visible=bool(is_visible),
    )
    
    return Response({
        'success': True,
        'batch': {
            'id': batch.id,
            'batch_id': batch.batch_id,
            'title': batch.title,
            'date': batch.date.isoformat(),
            'deadline': batch.deadline.isoformat(),
            'status': batch.status,
            'is_visible': batch.is_visible,
        }
    }, status=status.HTTP_201_CREATED)


@api_view(['PUT', 'PATCH'])
def admin_update_batch(request, batch_id):
    """
    PUT/PATCH /api/admin/batches/<batch_id>/
    Update batch.
    Query param: ?username=admin (wajib)
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Validasi admin via query parameter (konsisten dengan endpoint lain)
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        batch = Batch.objects.get(id=batch_id)
    except Batch.DoesNotExist:
        return Response({'error': 'Batch tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
    
    # Handle both DRF request.data and raw JSON
    if hasattr(request, 'data'):
        data = request.data
    else:
        import json
        try:
            data = json.loads(request.body) if request.body else {}
        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON body.'}, status=status.HTTP_400_BAD_REQUEST)
    
    logger.info(f"Updating batch {batch_id}, received data: {data}")
    
    # Validate required fields if provided
    if 'title' in data and (not data['title'] or (isinstance(data['title'], str) and not data['title'].strip())):
        return Response({'error': 'Judul batch tidak boleh kosong.'}, status=status.HTTP_400_BAD_REQUEST)
    
    if 'title' in data:
        batch.title = data['title'].strip()
    if 'date' in data:
        try:
            from datetime import datetime
            batch.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format tanggal tidak valid.'}, status=status.HTTP_400_BAD_REQUEST)
    if 'deadline' in data:
        try:
            from datetime import datetime
            batch.deadline = datetime.strptime(data['deadline'], '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format deadline tidak valid.'}, status=status.HTTP_400_BAD_REQUEST)
    if 'status' in data:
        batch.status = data['status']
    if 'description' in data:
        batch.description = data['description'].strip()
    
    # Handle is_visible explicitly
    if 'is_visible' in data:
        is_visible_raw = data['is_visible']
        if isinstance(is_visible_raw, str):
            batch.is_visible = is_visible_raw.lower() in ('true', '1', 'yes')
        else:
            batch.is_visible = bool(is_visible_raw)
        logger.info(f"Setting is_visible to {batch.is_visible} (from raw: {is_visible_raw})")
    
    try:
        batch.save()
        batch.refresh_from_db()
        logger.info(f"Batch {batch_id} updated successfully. is_visible={batch.is_visible}")
    except Exception as e:
        logger.error(f"Error saving batch {batch_id}: {e}", exc_info=True)
        return Response({'error': f'Gagal menyimpan batch: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'success': True,
        'batch': {
            'id': batch.id,
            'batch_id': batch.batch_id,
            'title': batch.title,
            'date': batch.date.isoformat(),
            'deadline': batch.deadline.isoformat(),
            'status': batch.status,
            'is_visible': batch.is_visible,
        }
    }, status=status.HTTP_200_OK)


@csrf_exempt
@require_http_methods(['DELETE'])
def admin_delete_batch(request, batch_id):
    """
    DELETE /api/admin/batches/<batch_id>/delete/
    Hapus batch.
    Query params: ?username=admin (wajib)
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"=== admin_delete_batch called for batch_id={batch_id} ===")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Query params: {dict(request.GET)}")
    
    # Validasi admin dari username query parameter
    username = request.GET.get('username', '').strip()
    logger.info(f"Username from query: '{username}'")
    
    if not username:
        logger.warning("No username provided in query params")
        return JsonResponse(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=400
        )
    
    try:
        admin_user = User.objects.get(username=username)
        logger.info(f"Admin user found: {admin_user.username}, is_staff: {admin_user.is_staff}")
        if not admin_user.is_staff:
            logger.warning(f"User {username} is not staff")
            return JsonResponse(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=403
            )
    except User.DoesNotExist:
        logger.error(f"Admin user not found: {username}")
        return JsonResponse(
            {'error': 'Admin user tidak ditemukan'},
            status=404
        )
    
    try:
        batch = Batch.objects.get(id=batch_id)
        logger.info(f"Batch found: {batch.batch_id}")
    except Batch.DoesNotExist:
        logger.error(f"Batch not found: {batch_id}")
        return JsonResponse({'error': 'Batch tidak ditemukan.'}, status=404)
    
    batch.delete()
    logger.info(f"Batch {batch_id} deleted successfully")
    return JsonResponse({'success': True}, status=200)


@api_view(['GET'])
def list_batches(request):
    """
    GET /api/batches/
    List semua batches yang visible (is_visible=True) untuk public/dashboard tryout.
    """
    try:
        batches = Batch.objects.filter(is_visible=True).order_by('-date', '-created_at')
        data = []
        for batch in batches:
            try:
                # Calculate duration from subtests (default 90 minutes if no subtests)
                duration = 90  # Default duration
                
                # Get subtests_info for this batch
                subtests_data = []
                try:
                    # Use select_related to avoid N+1 queries
                    soal_list = batch.soal_set.select_related('subtest').all()
                    for soal in soal_list:
                        if soal.subtest:
                            subtest_info = {
                                'code': soal.subtest.code,
                                'title': soal.subtest.nama_subtest,
                            }
                            # Check if not already in list (by code)
                            if not any(s['code'] == subtest_info['code'] for s in subtests_data):
                                subtests_data.append(subtest_info)
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Error getting subtests_info for batch {batch.batch_id}: {e}", exc_info=True)
                    subtests_data = []
                
                data.append({
                    'id': batch.id,
                    'batch_id': batch.batch_id,
                    'title': batch.title,
                    'date': batch.date.isoformat(),
                    'date_display': batch.date.strftime('%d %B %Y'),
                    'deadline': batch.deadline.isoformat(),
                    'deadline_display': batch.deadline.strftime('%d %B %Y'),
                    'status': batch.status,
                    'description': batch.description or '',
                    'duration': duration,
                    'subtests_info': subtests_data,
                })
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error processing batch {batch.batch_id}: {e}", exc_info=True)
                # Still add batch with minimal data
                data.append({
                    'id': batch.id,
                    'batch_id': batch.batch_id,
                    'title': batch.title,
                    'date': batch.date.isoformat(),
                    'date_display': batch.date.strftime('%d %B %Y'),
                    'deadline': batch.deadline.isoformat(),
                    'deadline_display': batch.deadline.strftime('%d %B %Y'),
                    'status': batch.status,
                    'description': batch.description or '',
                    'duration': 90,
                    'subtests_info': [],
                })
        
        return Response(data, status=status.HTTP_200_OK)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in list_batches: {e}", exc_info=True)
        return Response(
            {'error': 'Gagal mengambil data batch', 'detail': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def riwayat_nilai(request, username):
    """
    GET /api/riwayat-nilai/{username}/
    
    Ambil riwayat nilai tryout per subtest untuk user tertentu.
    """
    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    hasil_list = HasilTryout.objects.filter(user=user).select_related('subtest').order_by('-waktu_selesai', '-created_at')
    
    data = []
    for hasil in hasil_list:
        # Pastikan skor dalam range 0-100
        skor_value = float(hasil.skor) if hasil.skor is not None else 0.0
        if skor_value < 0:
            skor_value = 0.0
        elif skor_value > 100:
            skor_value = 100.0
        
        data.append({
            'id': hasil.id,
            'batch_id': hasil.batch_id,
            'subtest_code': hasil.subtest.code,
            'subtest_nama': hasil.subtest.nama_subtest,
            'jumlah_benar': hasil.jumlah_benar,
            'jumlah_salah': hasil.jumlah_salah,
            'jumlah_kosong': hasil.jumlah_kosong,
            'skor': round(skor_value, 2),  # Bulatkan ke 2 desimal
            'waktu_selesai': hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None,
            'tanggal': hasil.waktu_selesai.strftime('%Y-%m-%d') if hasil.waktu_selesai else hasil.created_at.strftime('%Y-%m-%d'),
        })
    
    return Response(data, status=status.HTTP_200_OK)
